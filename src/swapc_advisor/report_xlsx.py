"""Excel workbook generation with openpyxl.

Sheets: Summary, Ranked, Exchange Economics, POR Baselines, Disqualified,
Threat Reference, Provenance, Retrieved Context. Arial throughout.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import Recommendation, ReportError, ScoredCandidate

HEADER_FILL: Final[PatternFill] = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT: Final[Font] = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT: Final[Font] = Font(name="Arial", size=10)
BOLD_FONT: Final[Font] = Font(name="Arial", bold=True, size=10)
GOOD_FILL: Final[PatternFill] = PatternFill("solid", fgColor="E3F0E3")
BAD_FILL: Final[PatternFill] = PatternFill("solid", fgColor="FBE4E4")
WRAP: Final[Alignment] = Alignment(wrap_text=True, vertical="top")
CURRENCY_FMT: Final[str] = "$#,##0"


def _write_header(sheet: Worksheet, row: int, headers: tuple[str, ...]) -> None:
    """Write one styled header row."""
    for col, text in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def _set_widths(sheet: Worksheet, widths: tuple[int, ...]) -> None:
    """Apply column widths in order."""
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _cell(sheet: Worksheet, row: int, col: int, value: object) -> None:
    """Write one body cell with standard font and wrapping."""
    target = sheet.cell(row=row, column=col, value=value)
    target.font = BODY_FONT
    target.alignment = WRAP


def _money(sheet: Worksheet, row: int, col: int, value: float) -> None:
    """Write one currency-formatted body cell."""
    target = sheet.cell(row=row, column=col, value=value)
    target.font = BODY_FONT
    target.number_format = CURRENCY_FMT
    target.alignment = WRAP


def _pk_display(item: object) -> str:
    """Render per-group Pk as a readable string for a cell."""
    pk: dict[int, float] = getattr(item, "single_shot_pk", {})
    if not pk:
        return "n/a (enabler)"
    return ", ".join(f"G{g}: {v:.2f}" for g, v in sorted(pk.items()))


def _summary_sheet(book: Workbook, rec: Recommendation, disclaimer: str) -> None:
    """Sheet 1: query inputs, posture, BLUF, and disclaimer."""
    sheet: Worksheet = book.active
    sheet.title = "Summary"
    _set_widths(sheet, (30, 95))
    stamp: str = datetime.now(timezone.utc).strftime("%d %b %Y %H:%MZ")
    q = rec.query
    tiers: str = ", ".join(f"{k}: {v}" for k, v in sorted(rec.tier_distribution.items()))
    pairs: tuple[tuple[str, str], ...] = (
        ("Report", "SWaP-C Technology Down-Selection"),
        ("Classification", "UNCLASSIFIED // OPEN-SOURCE ESTIMATES"),
        ("Generated", stamp),
        ("Mission thread", rec.thread.name),
        ("AOR", f"{rec.aor.name} ({rec.aor.aor_id})"),
        ("Cost per unit threshold", f"${q.cost_threshold_usd:,.0f}"),
        ("Flight time requirement", f"{q.flight_time_min:.0f} min"),
        ("Distance requirement", f"{q.distance_km:.0f} km"),
        ("Scoring posture", f"{rec.posture.name}"),
        ("Posture intent", rec.posture.intent),
        ("Cost tier ceiling", q.max_cost_tier),
        ("Reference threat cost", f"${rec.aor.median_threat_cost_usd:,.0f}"),
        ("Exchange ratio target", f"{rec.thread.exchange_ratio_target:.0f}:1"),
        ("Assumed salvo size", f"{rec.thread.typical_salvo_size} threats"),
        ("Passing candidates by tier", tiers if tiers else "none"),
        ("BLUF", rec.summary),
        ("AOR threat profile", rec.aor.threat_profile),
        ("AOR exchange context", rec.aor.exchange_context),
        ("AOR operating notes", rec.aor.operating_notes),
        ("Disclaimer", disclaimer),
    )
    for row, (label, value) in enumerate(pairs, start=1):
        label_cell = sheet.cell(row=row, column=1, value=label)
        label_cell.font = BOLD_FONT
        label_cell.alignment = WRAP
        _cell(sheet, row, 2, value)


def _ranked_row(sheet: Worksheet, row: int, rank: int, cand: ScoredCandidate) -> None:
    """Write one ranked-candidate row across all columns."""
    e = cand.equipment
    cls = cand.classification
    _cell(sheet, row, 1, rank)
    _cell(sheet, row, 2, e.name)
    _cell(sheet, row, 3, e.vendor)
    _cell(sheet, row, 4, e.vendor_maturity)
    _cell(sheet, row, 5, f"{cls.cost_tier} {cls.cost_tier_name}")
    _cell(sheet, row, 6, f"{cls.swap_tier} {cls.swap_tier_name}")
    _cell(sheet, row, 7, round(cand.total_score, 3))
    _money(sheet, row, 8, e.swap_c.unit_cost_usd)
    _cell(sheet, row, 9, e.swap_c.weight_lb)
    _cell(sheet, row, 10, e.swap_c.power_w)
    _cell(sheet, row, 11, e.flight_time_min if e.flight_time_min > 0 else "N/A")
    _cell(sheet, row, 12, e.effective_range_km)
    _cell(sheet, row, 13, ", ".join(str(g) for g in e.target_groups))
    _cell(sheet, row, 14, e.units_per_month)
    _cell(sheet, row, 15, e.evidence_grade)
    _cell(sheet, row, 16, "Yes" if e.gps_denied_capable else "No")
    _cell(sheet, row, 17, "Yes" if e.ew_resilient else "No")
    _cell(sheet, row, 18, " ".join(cand.rationale))


def _ranked_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet 2: full ranked candidate table with tiers and SWaP detail."""
    sheet: Worksheet = book.create_sheet("Ranked")
    headers: tuple[str, ...] = (
        "Rank",
        "System",
        "Vendor",
        "Vendor Maturity",
        "Cost Tier",
        "SWaP Tier",
        "Score",
        "Unit Cost (USD)",
        "Weight (lb)",
        "Power (W)",
        "Flight Time (min)",
        "Range (km)",
        "Target Groups",
        "Production/Month",
        "Evidence Grade",
        "GPS-Denied",
        "EW Resilient",
        "Rationale",
    )
    _write_header(sheet, 1, headers)
    _set_widths(
        sheet,
        (6, 28, 20, 15, 20, 20, 8, 14, 11, 10, 13, 11, 13, 15, 15, 11, 12, 80),
    )
    for rank, cand in enumerate(rec.ranked, start=1):
        _ranked_row(sheet, rank + 1, rank, cand)
    sheet.freeze_panes = "B2"


def _exchange_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet 3: cost-exchange economics with favorability shading."""
    sheet: Worksheet = book.create_sheet("Exchange Economics")
    headers: tuple[str, ...] = (
        "System",
        "Cost Tier",
        "Unit Cost (USD)",
        "Uses/Unit",
        "Pk by Group",
        "Rounds/Defeat",
        "Cost per Defeat (USD)",
        "Threat Cost (USD)",
        "Exchange Ratio",
        "Favorable",
        "Magazine Rounds",
        "Magazine Cost (USD)",
        "Replenish (days)",
        "Assessment",
    )
    _write_header(sheet, 1, headers)
    _set_widths(sheet, (28, 12, 14, 10, 12, 12, 18, 15, 13, 11, 14, 17, 13, 70))
    combined: tuple[ScoredCandidate, ...] = rec.ranked + rec.enablers + rec.baselines
    for row, cand in enumerate(combined, start=2):
        e, ex = cand.equipment, cand.exchange
        _cell(sheet, row, 1, e.name)
        _cell(sheet, row, 2, cand.classification.cost_tier)
        _money(sheet, row, 3, e.swap_c.unit_cost_usd)
        _cell(sheet, row, 4, e.uses_per_unit)
        _cell(sheet, row, 5, _pk_display(e))
        _cell(sheet, row, 6, ex.rounds_per_defeat)
        _money(sheet, row, 7, ex.cost_per_defeat_usd)
        _money(sheet, row, 8, ex.threat_cost_usd)
        _cell(sheet, row, 9, ex.exchange_ratio)
        verdict = sheet.cell(row=row, column=10, value="Yes" if ex.favorable else "No")
        verdict.font = BODY_FONT
        verdict.fill = GOOD_FILL if ex.favorable else BAD_FILL
        _cell(sheet, row, 11, ex.magazine_rounds)
        _money(sheet, row, 12, ex.magazine_cost_usd)
        _cell(sheet, row, 13, ex.replenish_days)
        _cell(sheet, row, 14, ex.notes)
    sheet.freeze_panes = "B2"


def _enabler_sheet(book: Workbook, rec: Recommendation) -> None:
    """Detection and cueing layer, separate from effectors."""
    sheet: Worksheet = book.create_sheet("Enablers")
    _write_header(
        sheet,
        1,
        (
            "System",
            "Vendor",
            "Cost Tier",
            "SWaP Tier",
            "Unit Cost (USD)",
            "Range (km)",
            "Weight (lb)",
            "Power (W)",
            "Evidence Grade",
            "Note",
        ),
    )
    _set_widths(sheet, (30, 22, 12, 20, 15, 11, 11, 10, 16, 85))
    for row, cand in enumerate(rec.enablers, start=2):
        e = cand.equipment
        _cell(sheet, row, 1, e.name)
        _cell(sheet, row, 2, e.vendor)
        _cell(sheet, row, 3, cand.classification.cost_tier)
        _cell(
            sheet, row, 4, f"{cand.classification.swap_tier} {cand.classification.swap_tier_name}"
        )
        _money(sheet, row, 5, e.swap_c.unit_cost_usd)
        _cell(sheet, row, 6, e.effective_range_km)
        _cell(sheet, row, 7, e.swap_c.weight_lb)
        _cell(sheet, row, 8, e.swap_c.power_w)
        _cell(sheet, row, 9, e.evidence_grade)
        _cell(sheet, row, 10, e.notes)


def _baseline_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet 4: POR comparators and the delta against the top recommendation."""
    sheet: Worksheet = book.create_sheet("POR Baselines")
    _write_header(
        sheet,
        1,
        (
            "System",
            "Vendor",
            "Cost Tier",
            "Unit Cost (USD)",
            "Cost per Defeat (USD)",
            "Exchange Ratio",
            "Delta vs Top Pick",
            "Note",
        ),
    )
    _set_widths(sheet, (28, 20, 12, 15, 20, 14, 18, 85))
    top_cost: float = rec.ranked[0].exchange.cost_per_defeat_usd if rec.ranked else 0.0
    for row, cand in enumerate(rec.baselines, start=2):
        e, ex = cand.equipment, cand.exchange
        delta: str = "n/a"
        if top_cost > 0 and ex.cost_per_defeat_usd > 0:
            delta = f"{ex.cost_per_defeat_usd / top_cost:.1f}x costlier per defeat"
        _cell(sheet, row, 1, e.name)
        _cell(sheet, row, 2, e.vendor)
        _cell(sheet, row, 3, cand.classification.cost_tier)
        _money(sheet, row, 4, e.swap_c.unit_cost_usd)
        _money(sheet, row, 5, ex.cost_per_defeat_usd)
        _cell(sheet, row, 6, ex.exchange_ratio)
        _cell(sheet, row, 7, delta)
        _cell(sheet, row, 8, e.notes)


def _rejected_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet 5: disqualified systems with explicit gate failures."""
    sheet: Worksheet = book.create_sheet("Disqualified")
    _write_header(
        sheet,
        1,
        ("System", "Vendor", "Cost Tier", "Unit Cost (USD)", "Disqualifiers"),
    )
    _set_widths(sheet, (28, 20, 12, 15, 85))
    for row, cand in enumerate(rec.rejected, start=2):
        _cell(sheet, row, 1, cand.equipment.name)
        _cell(sheet, row, 2, cand.equipment.vendor)
        _cell(sheet, row, 3, cand.classification.cost_tier)
        _money(sheet, row, 4, cand.equipment.swap_c.unit_cost_usd)
        _cell(sheet, row, 5, "; ".join(cand.disqualifiers))


def _reference_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet 6: UAS group threat cost and affordable defeat ceilings."""
    sheet: Worksheet = book.create_sheet("Threat Reference")
    _write_header(
        sheet,
        1,
        (
            "Group",
            "Max Weight (lb)",
            "Endurance (min)",
            "Range (km)",
            "Median Threat Cost (USD)",
            "Affordable Defeat Ceiling (USD)",
            "Representative Threats",
            "Notes",
        ),
    )
    _set_widths(sheet, (22, 14, 16, 14, 22, 26, 36, 85))
    for row, group in enumerate(rec.relevant_groups, start=2):
        _cell(sheet, row, 1, group.name)
        _cell(sheet, row, 2, group.max_weight_lb)
        _cell(sheet, row, 3, f"{group.endurance_min[0]:.0f}-{group.endurance_min[1]:.0f}")
        _cell(sheet, row, 4, f"{group.range_km[0]:.0f}-{group.range_km[1]:.0f}")
        _money(sheet, row, 5, group.median_threat_cost_usd)
        _money(sheet, row, 6, group.affordable_defeat_ceiling_usd)
        _cell(sheet, row, 7, ", ".join(group.representative_threats))
        _cell(sheet, row, 8, group.notes)


def _provenance_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet 7: data confidence and watch items for every candidate."""
    sheet: Worksheet = book.create_sheet("Provenance")
    _write_header(
        sheet,
        1,
        (
            "System",
            "Cost Confidence",
            "As Of",
            "Evidence Grade",
            "Vendor Maturity",
            "Source Note",
            "Watch Items",
        ),
    )
    _set_widths(sheet, (28, 18, 10, 16, 15, 70, 80))
    combined: tuple[ScoredCandidate, ...] = rec.ranked + rec.enablers + rec.baselines + rec.rejected
    for row, cand in enumerate(combined, start=2):
        e = cand.equipment
        confidence = sheet.cell(row=row, column=2, value=e.cost_confidence)
        confidence.font = BODY_FONT
        confidence.alignment = WRAP
        if e.cost_confidence == "order_of_magnitude":
            confidence.fill = BAD_FILL
        elif e.cost_confidence == "published":
            confidence.fill = GOOD_FILL
        _cell(sheet, row, 1, e.name)
        _cell(sheet, row, 3, e.as_of)
        _cell(sheet, row, 4, e.evidence_grade)
        _cell(sheet, row, 5, e.vendor_maturity)
        _cell(sheet, row, 6, e.source_note)
        _cell(sheet, row, 7, " | ".join(cand.watch_items))
    sheet.freeze_panes = "B2"


def _sensitivity_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet: rank stability under the 9-scenario perturbation grid."""
    sheet: Worksheet = book.create_sheet("Sensitivity")
    _write_header(sheet, 1, ("System", "Best Rank", "Worst Rank", "Verdict"))
    _set_widths(sheet, (30, 12, 12, 40))
    if rec.sensitivity is None:
        _cell(sheet, 2, 1, "No effectors passed; sensitivity not computed.")
        return
    _cell(sheet, 2, 1, rec.sensitivity.notes)
    row: int = 4
    for cand in rec.ranked:
        low, high = rec.sensitivity.rank_ranges.get(cand.equipment.equipment_id, (0, 0))
        _cell(sheet, row, 1, cand.equipment.name)
        _cell(sheet, row, 2, low)
        _cell(sheet, row, 3, high)
        verdict_cell = sheet.cell(
            row=row,
            column=4,
            value="Stable" if low == high else f"Varies rank {low}-{high}",
        )
        verdict_cell.font = BODY_FONT
        verdict_cell.fill = GOOD_FILL if low == high else BAD_FILL
        row += 1


def _architecture_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet: layered defense proposal."""
    sheet: Worksheet = book.create_sheet("Architecture")
    _write_header(
        sheet,
        1,
        ("Layer", "Band (km)", "System", "Effective Pk", "Magazine Cost (USD)"),
    )
    _set_widths(sheet, (12, 12, 32, 12, 20))
    if rec.architecture is None:
        _cell(sheet, 2, 1, "No effectors passed; architecture not proposed.")
        return
    arch = rec.architecture
    row: int = 2
    for layer in arch.layers:
        _cell(sheet, row, 1, layer.band)
        _cell(sheet, row, 2, f"{layer.band_range_km[0]:.0f}-{layer.band_range_km[1]:.0f}")
        _cell(sheet, row, 3, layer.system_name)
        _cell(sheet, row, 4, layer.effective_pk)
        _money(sheet, row, 5, layer.magazine_cost_usd)
        row += 1
    _cell(sheet, row, 1, "Sensor")
    _cell(sheet, row, 3, arch.sensor_name)
    _money(sheet, row, 5, arch.sensor_cost_usd)
    row += 2
    _cell(sheet, row, 1, "Total magazine cost")
    _money(sheet, row, 5, arch.total_magazine_cost_usd)
    row += 1
    _cell(sheet, row, 1, "Leakage probability")
    _cell(sheet, row, 5, arch.leakage_probability)
    row += 1
    _cell(sheet, row, 1, arch.notes)


def _context_sheet(book: Workbook, rec: Recommendation) -> None:
    """Sheet 8: retrieved RAG passages with section tags and scores."""
    sheet: Worksheet = book.create_sheet("Retrieved Context")
    _write_header(sheet, 1, ("Doc ID", "Title", "Section", "Relevance", "Passage"))
    _set_widths(sheet, (22, 30, 14, 11, 115))
    for row, doc in enumerate(rec.retrieved_context, start=2):
        _cell(sheet, row, 1, doc.doc_id)
        _cell(sheet, row, 2, doc.title)
        _cell(sheet, row, 3, doc.section)
        _cell(sheet, row, 4, round(doc.score, 2))
        _cell(sheet, row, 5, doc.text)


def write_xlsx(rec: Recommendation, out_path: Path, disclaimer: str) -> Path:
    """Render the full recommendation workbook to an xlsx file."""
    book: Workbook = Workbook()
    _summary_sheet(book, rec, disclaimer)
    _ranked_sheet(book, rec)
    _exchange_sheet(book, rec)
    _enabler_sheet(book, rec)
    _baseline_sheet(book, rec)
    _rejected_sheet(book, rec)
    _reference_sheet(book, rec)
    _sensitivity_sheet(book, rec)
    _architecture_sheet(book, rec)
    _provenance_sheet(book, rec)
    _context_sheet(book, rec)
    try:
        book.save(str(out_path))
    except OSError as exc:
        raise ReportError(f"Excel generation failed: {exc}") from exc
    return out_path
